import asyncio
import logging
from aiogram import Bot, Dispatcher, types
from aiogram.contrib.middlewares.logging import LoggingMiddleware
from aiogram.types import ParseMode
from datetime import datetime
from aiogram import executor
from openpyxl import Workbook, load_workbook
import os

logging.basicConfig(level=logging.INFO)


bot = Bot(token="")
dp = Dispatcher(bot)
dp.middleware.setup(LoggingMiddleware())


EXCEL_FILE_PATH = "data1.xlsx"



def append_to_excel(username, date):

    if os.path.exists(EXCEL_FILE_PATH):

        wb = load_workbook(EXCEL_FILE_PATH)

        sheet = wb.active
    else:

        wb = Workbook()

        sheet = wb.active

        sheet.append(["User", "Time"])


    sheet.append([username, date])


    wb.save(EXCEL_FILE_PATH)



@dp.message_handler(commands=['start'])
async def process_start_command(message: types.Message):

    username = message.from_user.username

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    append_to_excel(username, now)

    await message.reply(f"Salom, {username}")



async def main():

    await dp.start_polling()

    await asyncio.gather(*asyncio.all_tasks())


if __name__ == '__main__':
    asyncio.run(main())
